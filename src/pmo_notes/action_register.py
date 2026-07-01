"""Registre d'actions inter-réunions.

Le cœur du métier PMO : agréger, au fil des réunions, les « actions à venir »
extraites des synthèses dans un registre cumulatif (Excel et/ou CSV), avec le
responsable, l'échéance, le statut et la réunion d'origine.

L'extraction des actions depuis le Markdown de la synthèse est *pure* (donc
testable sans dépendance) ; l'écriture Excel s'appuie sur `openpyxl` (importé
paresseusement). Le CSV, lui, est toujours produit (bibliothèque standard).
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

#: En-têtes de colonnes du registre.
HEADERS = ["Date", "Réunion", "Responsable", "Action", "Échéance", "Statut", "Source"]

#: Statut par défaut d'une action nouvellement extraite.
DEFAULT_STATUS = "À faire"

#: Statuts possibles d'une action (pour la vue du registre).
STATUSES = ["À faire", "En cours", "Fait", "Annulé"]

# Puce « - **Responsable** — action (échéance) » (le séparateur peut être — ou -).
_BULLET_RE = re.compile(r"^\s*[-*]\s+(?P<body>.+?)\s*$")
_RESP_RE = re.compile(r"^\*\*(?P<resp>.+?)\*\*\s*[—–-]\s*(?P<rest>.+)$")
_ECHEANCE_RE = re.compile(r"\((?P<ech>[^()]*)\)\s*$")
_HEADING_RE = re.compile(r"^##\s+(?P<title>.+?)\s*$")


class RegisterError(RuntimeError):
    """Erreur d'écriture du registre d'actions."""


@dataclass
class ActionItem:
    """Une action à suivre, telle qu'inscrite au registre."""

    date: str
    reunion: str
    responsable: str
    action: str
    echeance: str = ""
    statut: str = DEFAULT_STATUS
    source: str = ""

    def as_row(self) -> list[str]:
        return [
            self.date,
            self.reunion,
            self.responsable,
            self.action,
            self.echeance,
            self.statut,
            self.source,
        ]


def _strip_bold(text: str) -> str:
    return text.replace("**", "").strip()


def extract_actions(
    synthesis: str,
    *,
    meeting: str,
    date: str,
    source: str = "",
) -> list[ActionItem]:
    """Extrait les actions de la section « Actions à venir » d'une synthèse."""
    lines = synthesis.splitlines()
    # Localise la section dont le titre contient « action ».
    start = None
    for i, line in enumerate(lines):
        heading = _HEADING_RE.match(line)
        if heading and "action" in heading.group("title").lower():
            start = i + 1
            break
    if start is None:
        return []

    items: list[ActionItem] = []
    for line in lines[start:]:
        if _HEADING_RE.match(line):  # section suivante : on s'arrête
            break
        bullet = _BULLET_RE.match(line)
        if not bullet:
            continue
        body = bullet.group("body").strip()

        resp_match = _RESP_RE.match(body)
        if resp_match:
            responsable = _strip_bold(resp_match.group("resp"))
            rest = resp_match.group("rest").strip()
        else:
            responsable = "à confirmer"
            rest = body

        echeance = ""
        ech_match = _ECHEANCE_RE.search(rest)
        if ech_match:
            echeance = ech_match.group("ech").strip()
            rest = rest[: ech_match.start()].strip()

        action = _strip_bold(rest)
        if not action:
            continue
        items.append(
            ActionItem(
                date=date,
                reunion=meeting,
                responsable=responsable,
                action=action,
                echeance=echeance,
                source=source,
            )
        )
    return items


def append_to_csv(path: Path, items: list[ActionItem]) -> Path:
    """Ajoute des actions à un registre CSV (créé avec ses en-têtes si absent).

    Séparateur « ; » et BOM UTF-8 (écrit une seule fois, à la création) pour une
    ouverture correcte dans Excel, notamment en français. Le BOM est géré
    manuellement : `utf-8-sig` en mode « append » réémettrait un BOM au milieu
    du fichier à chaque ajout.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    is_new = not path.exists()
    with path.open("a", newline="", encoding="utf-8") as handle:
        if is_new:
            handle.write("﻿")  # BOM, une seule fois
        writer = csv.writer(handle, delimiter=";")
        if is_new:
            writer.writerow(HEADERS)
        for item in items:
            writer.writerow(item.as_row())
    return path


def write_csv(path: Path, items: list[ActionItem]) -> Path:
    """Réécrit intégralement le registre CSV (en-têtes + lignes)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        handle.write("﻿")
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(HEADERS)
        for item in items:
            writer.writerow(item.as_row())
    return path


def read_csv(path: Path) -> list[ActionItem]:
    """Lit un registre CSV en liste d'actions (vide si le fichier est absent)."""
    path = Path(path)
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    if not rows:
        return []
    start = 1 if rows[0] and rows[0][0] == HEADERS[0] else 0
    items: list[ActionItem] = []
    for row in rows[start:]:
        if not any(cell.strip() for cell in row):
            continue
        vals = (row + [""] * len(HEADERS))[: len(HEADERS)]
        items.append(
            ActionItem(
                date=vals[0],
                reunion=vals[1],
                responsable=vals[2],
                action=vals[3],
                echeance=vals[4],
                statut=vals[5] or DEFAULT_STATUS,
                source=vals[6],
            )
        )
    return items


def append_to_xlsx(path: Path, items: list[ActionItem]) -> Path:
    """Ajoute des actions à un classeur Excel (créé avec ses en-têtes si absent)."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover - dépendance optionnelle
        raise RegisterError(
            "Le paquet « openpyxl » est requis pour le registre au format Excel "
            "(pip install openpyxl)."
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Actions"
        worksheet.append(HEADERS)
    for item in items:
        worksheet.append(item.as_row())
    workbook.save(path)
    return path


def write_xlsx(path: Path, items: list[ActionItem]) -> Path:
    """Réécrit intégralement le classeur Excel (en-têtes + lignes)."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dépendance optionnelle
        raise RegisterError(
            "Le paquet « openpyxl » est requis pour le registre au format Excel "
            "(pip install openpyxl)."
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Actions"
    worksheet.append(HEADERS)
    for item in items:
        worksheet.append(item.as_row())
    workbook.save(path)
    return path


def update_register(output_dir: Path, items: list[ActionItem]) -> list[Path]:
    """Met à jour le registre (CSV toujours ; Excel si `openpyxl` est présent)."""
    output_dir = Path(output_dir)
    if not items:
        return []
    paths = [append_to_csv(output_dir / "registre_actions.csv", items)]
    try:
        paths.append(append_to_xlsx(output_dir / "registre_actions.xlsx", items))
    except RegisterError:
        pass  # openpyxl absent : le CSV suffit
    return paths


def csv_path(output_dir: Path) -> Path:
    return Path(output_dir) / "registre_actions.csv"


def read_register(output_dir: Path) -> list[ActionItem]:
    """Lit le registre d'actions (depuis le CSV, source de référence)."""
    return read_csv(csv_path(output_dir))


def write_register(output_dir: Path, items: list[ActionItem]) -> list[Path]:
    """Réécrit intégralement le registre (CSV toujours ; Excel si disponible)."""
    output_dir = Path(output_dir)
    paths = [write_csv(csv_path(output_dir), items)]
    try:
        paths.append(write_xlsx(output_dir / "registre_actions.xlsx", items))
    except RegisterError:
        pass  # openpyxl absent : le CSV suffit
    return paths


def filter_items(
    items: list[ActionItem],
    responsable: Optional[str] = None,
    statut: Optional[str] = None,
) -> list[ActionItem]:
    """Filtre les actions par responsable et/ou statut (None = pas de filtre)."""
    result = items
    if responsable:
        result = [i for i in result if i.responsable == responsable]
    if statut:
        result = [i for i in result if i.statut == statut]
    return list(result)


__all__ = [
    "RegisterError",
    "ActionItem",
    "HEADERS",
    "DEFAULT_STATUS",
    "STATUSES",
    "extract_actions",
    "append_to_csv",
    "append_to_xlsx",
    "update_register",
    "write_csv",
    "read_csv",
    "write_xlsx",
    "csv_path",
    "read_register",
    "write_register",
    "filter_items",
]
