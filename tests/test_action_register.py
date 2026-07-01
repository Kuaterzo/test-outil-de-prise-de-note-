import pytest

from pmo_notes.action_register import (
    ActionItem,
    append_to_csv,
    append_to_xlsx,
    extract_actions,
    filter_items,
    read_csv,
    read_register,
    update_register,
    write_csv,
    write_register,
)


def _items():
    return [
        ActionItem("2026-06-26", "COPIL", "Alice", "Préparer le plan", "lundi", "À faire", "s.md"),
        ActionItem("2026-06-26", "COPIL", "Bob", "Valider le budget", "", "Fait", "s.md"),
    ]

SAMPLE = (
    "## Introduction\n"
    "Réunion de cadrage.\n\n"
    "## Actions à venir\n"
    "- **Alice** — rédiger le plan (vendredi)\n"
    "- **Bob** — valider le budget\n"
    "- relancer le prestataire\n\n"
    "## Conclusion\n"
    "Fin.\n"
)


def test_extract_actions_parses_responsible_and_due():
    items = extract_actions(SAMPLE, meeting="COPIL", date="2026-06-26", source="s.md")
    assert len(items) == 3

    assert items[0].responsable == "Alice"
    assert items[0].action == "rédiger le plan"
    assert items[0].echeance == "vendredi"
    assert items[0].reunion == "COPIL"
    assert items[0].statut == "À faire"

    assert items[1].responsable == "Bob"
    assert items[1].echeance == ""

    # Puce sans responsable explicite
    assert items[2].responsable == "à confirmer"
    assert items[2].action == "relancer le prestataire"


def test_extract_actions_no_section_returns_empty():
    assert extract_actions("## Introduction\nrien", meeting="m", date="d") == []


def test_append_to_csv_creates_then_appends(tmp_path):
    path = tmp_path / "registre_actions.csv"
    items = extract_actions(SAMPLE, meeting="COPIL", date="2026-06-26")

    append_to_csv(path, items)
    lines = [l for l in path.read_text(encoding="utf-8-sig").splitlines() if l.strip()]
    assert lines[0].startswith("Date;Réunion;Responsable")
    assert len(lines) == 1 + 3  # en-tête + 3 actions

    append_to_csv(path, items[:1])
    lines = [l for l in path.read_text(encoding="utf-8-sig").splitlines() if l.strip()]
    assert len(lines) == 1 + 3 + 1  # en-tête conservé, 1 action ajoutée


def test_update_register_returns_paths(tmp_path):
    items = extract_actions(SAMPLE, meeting="COPIL", date="2026-06-26")
    paths = update_register(tmp_path, items)
    assert any(p.name == "registre_actions.csv" for p in paths)
    assert all(p.exists() for p in paths)


def test_update_register_empty_returns_nothing(tmp_path):
    assert update_register(tmp_path, []) == []


def test_csv_write_read_roundtrip(tmp_path):
    path = tmp_path / "registre_actions.csv"
    write_csv(path, _items())
    back = read_csv(path)
    assert [(i.responsable, i.action, i.statut) for i in back] == [
        ("Alice", "Préparer le plan", "À faire"),
        ("Bob", "Valider le budget", "Fait"),
    ]


def test_append_keeps_single_bom(tmp_path):
    path = tmp_path / "registre_actions.csv"
    append_to_csv(path, _items())
    append_to_csv(path, _items())  # 2e append : pas de BOM au milieu du fichier
    assert path.read_bytes().count(b"\xef\xbb\xbf") == 1
    assert len(read_csv(path)) == 4


def test_read_csv_missing_returns_empty(tmp_path):
    assert read_csv(tmp_path / "absent.csv") == []


def test_filter_items():
    items = _items()
    assert [i.responsable for i in filter_items(items, responsable="Alice")] == ["Alice"]
    assert [i.responsable for i in filter_items(items, statut="Fait")] == ["Bob"]
    assert filter_items(items, responsable="Alice", statut="Fait") == []
    assert len(filter_items(items)) == 2


def test_register_write_then_read_reflects_status_change(tmp_path):
    items = _items()
    write_register(tmp_path, items)
    # Modifie un statut puis réécrit.
    back = read_register(tmp_path)
    back[0].statut = "En cours"
    write_register(tmp_path, back)
    reread = read_register(tmp_path)
    assert reread[0].statut == "En cours"
    assert (tmp_path / "registre_actions.csv").exists()


def test_append_to_xlsx_appends_rows(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    path = tmp_path / "registre_actions.xlsx"
    items = extract_actions(SAMPLE, meeting="COPIL", date="2026-06-26")
    append_to_xlsx(path, items)
    append_to_xlsx(path, items[:1])

    worksheet = load_workbook(path).active
    # 1 en-tête + 3 + 1 actions = 5 lignes
    assert worksheet.max_row == 5
    assert [c.value for c in worksheet[1]] == [
        "Date", "Réunion", "Responsable", "Action", "Échéance", "Statut", "Source",
    ]
